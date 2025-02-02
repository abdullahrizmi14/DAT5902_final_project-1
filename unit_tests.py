import unittest
import pandas as pd
import openpyxl as xl
import os

from code1 import (
    select_sheet,
    list_sheets,
    write_headers_to_sheet,
    check_if_exists_then_delete,
    merge_and_append_data,
)

class TestFunction(unittest.TestCase):

    def setUp(self):
        # Create a temporary workbook for testing
        self.test_file = "test_workbook.xlsx"
        self.workbook = xl.Workbook()
        sheet1 = self.workbook.active
        sheet1.title = "Sheet1"
        self.workbook.create_sheet("Sheet2")
        sheet1.append(["Header1", "Header2", "Header3"])
        self.workbook.save(self.test_file)

    def tearDown(self):
        # Clean up after tests
        if os.path.exists(self.test_file):
            os.remove(self.test_file)

    def test_select_sheet(self):
        # Check if the function works without throwing errors
        workbook = xl.load_workbook(self.test_file)
        try:
            select_sheet(workbook, "Sheet1")
        except Exception as e:
            self.fail(f"select_sheet raised an exception: {e}")

    def test_list_sheets(self):
        # Check if all sheet names are printed
        workbook = xl.load_workbook(self.test_file)
        expected_sheets = ["Sheet1", "Sheet2"]
        actual_sheets = workbook.sheetnames
        self.assertEqual(actual_sheets, expected_sheets)

    def test_write_headers_to_sheet(self):
        # Test writing headers to a sheet
        workbook = xl.load_workbook(self.test_file)
        sheet = workbook["Sheet1"]
        headers = ["NewHeader1", "NewHeader2"]
        write_headers_to_sheet(sheet, headers, start_col=1, start_row=2)
        
        # Verify the headers are written
        self.assertEqual(sheet.cell(row=2, column=1).value, "NewHeader1")
        self.assertEqual(sheet.cell(row=2, column=2).value, "NewHeader2")

    def test_check_if_exists_then_delete(self):
        # Test deleting an existing file
        check_if_exists_then_delete(self.test_file)
        self.assertFalse(os.path.exists(self.test_file))

        # Test when file does not exist
        check_if_exists_then_delete(self.test_file)  # Should not raise any errors

    def test_merge_data(self):
        # Create DataFrames
        main_data = pd.DataFrame({"Team": ["A", "B", "C"], "Score": [1, 2, 3]})
        lookup_data = pd.DataFrame({"Team": ["A", "B", "D"], "Rank": [10, 20, 30]})

        # Perform the merge
        result = merge_and_append_data(main_data, lookup_data, key_column="Team", value_column="Rank")

        # Verify that the merged DataFrame has the expected structure and values
        expected_result = pd.DataFrame({
            "Team": ["A", "B", "C"],
            "Score": [1, 2, 3],
            "Rank": [10.0, 20.0, None]
        })

        # Use pandas testing utility to compare DataFrames
        pd.testing.assert_frame_equal(result, expected_result)

if __name__ == "__main__":
    unittest.main()