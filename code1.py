import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl as xl      # To install this library 'pip install openpyxl'
import os
from adjustText import adjust_text


## Functions Used ##
def select_sheet(workbook, name):      # Display the first few rows of the sheet from excel workbook
    """
    Display the first few rows of the specified sheet from an openpyxl workbook.
    
    Args:
        workbook: An openpyxl Workbook object.
        name: The name of the sheet to display.
    """
    if name in workbook.sheetnames:
        df = workbook[name]
        print("Data from " + name + ":")
        for row in df.iter_rows(min_row=1, max_row=5, values_only=True):  # Display the first few rows of the DataFrame
            print(row)    
    else:
        print(name + "sheet not found.")

def list_sheets(workbook):
    """
    Prints the names of all sheets in an Excel workbook.

    Args:
        workbook (openpyxl.Workbook): The Excel workbook object.

    Returns:
        None
    """
    sheetNames = workbook.sheetnames
    for name in sheetNames:
        print(name)

def write_headers_to_sheet(sheet, headers, start_col=1, start_row=1):
    """
    Writes a list of headers into a specified sheet starting from a given column and row.

    Args:
        sheet: The target worksheet (openpyxl worksheet object).
        headers: List of headers to write.
        start_col: Starting column to write headers (default 1).
        start_row: Row to write headers (default 1).
    """
    for col_index, header in enumerate(headers, start=start_col):
        print(f"Writing header '{header}' to column {col_index}")
        cell = sheet.cell(row=start_row, column=col_index)
        cell.value = header

def check_if_exists_then_delete(file_path):
    """
    Checks if a file exists at the specified path and deletes it if found. 
    If no file is found, it logs a message indicating its absence.

    Args:
        file_path (str): The path to the file.

    Returns:
        None
    """
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"Existing file'{file_path}' deleted.")
    else:
        print(f"No existing file found with name'{file_path}'.")


## Loading in original workbook ##
data_xl = xl.load_workbook('The_Data_Landscape_Project_Stats_Macro.xlsx')

## Checking sheets for original workbook ##
print('---------------------------------- Original columns -----------------------------------------------')
list_sheets(data_xl)

## Deleting Merged data sheet and creating a new one ##
data_xl.create_sheet('merged_data_python')
del data_xl['Merged Data']

## Copying in team names to new merged sheet ## 
final_data_sheet = data_xl['merged_data_python']

for i in range(1,34):
    for j in range(1,2):
        read_value = data_xl.worksheets[0].cell(row = i, column = j) 
        final_data_sheet.cell(row = i, column = j).value = read_value.value

## Checking sheets of new workbook ##
print('----------------------------------- New columns ----------------------------------------------------')
list_sheets(data_xl)               ## checking to make sure that new sheet and 1st column are implemented

## Save the workbook ##
file_name = 'transformation_workbook.xlsx'
check_if_exists_then_delete(file_name)
data_xl.save(file_name)
print(f"New file '{file_name}' created successfully.")


## Load new workbook for transformations and check head ##
tf = xl.load_workbook(file_name)
select_sheet(tf,'merged_data_python')    ## Checking head of sheet 


## Getting sheet names to make column headers ##
tf_shts = tf.sheetnames
headers = tf_shts[:-1]
last_sheet_name = tf_shts[-1]
last_sheet_tf = tf[last_sheet_name]

start_col = 2
start_row = 1

write_headers_to_sheet(last_sheet_tf, headers, start_col=2)

# Save
tf.save(file_name)


## Vlookup ##

# Loading Workbook and sheet names
main_wb = pd.read_excel(file_name, sheet_name=None)
sheetNames_wb = list(main_wb.keys())

# Loading last sheet and columns
main_data = pd.read_excel(file_name, sheet_name='merged_data_python')  # Main data
columns_md = list(main_data.columns)


for i, sheet_name in enumerate(sheetNames_wb):
    if i >= len(columns_md)-1:
        print("Not enough columns in main_data to update. Skipping remaining sheets.")
        break

    # Iterating through sheets
    lookup_data = pd.read_excel(file_name, sheet_name=sheet_name)  # Lookup data
    columns_ld = list(lookup_data.columns)

    # Merge on common key 'Team'
    merged_data = main_data.merge(lookup_data[columns_ld[:2]], on='Team', how='left')

    main_data[columns_md[i + 1]] = merged_data[columns_ld[1]]

# Save to CSV
main_data.to_csv('final_data.csv', index=False)
print("VLOOKUP operation completed successfully.")



#######################################################################################################################

## Analysis and Figures ##

## Loading Data ##
df = pd.read_csv('final_data.csv')
df

## Figure 1 'Rushing and Passing Play %'s' ##

## Storing data for Rushing Play % and Passing Play %
rushing_03 = df['Rushing Play % (2003)']
rushing_13 = df['Rushing Play % (2013)']
rushing_23 = df['Rushing Play % (2023)']

passing_03 = df['Passing Play % (2003)']
passing_13 = df['Passing Play % (2013)']
passing_23 = df['Passing Play % (2023)']

# Create a figure
plt.figure(figsize=(10, 6))

# Plot KDE for Rushing Play %
sns.kdeplot(rushing_03, shade=True, label='Rushing (2003)', color='blue', bw_adjust=1.2)
sns.kdeplot(rushing_13, shade=True, label='Rushing (2013)', color='green', bw_adjust=1.2)
sns.kdeplot(rushing_23, shade=True, label='Rushing (2023)', color='red', bw_adjust=1.2)

# Plot KDE for Passing Play %
sns.kdeplot(passing_03, shade=True, label='Passing (2003)', color='purple', bw_adjust=1.2)
sns.kdeplot(passing_13, shade=True, label='Passing (2013)', color='orange', bw_adjust=1.2)
sns.kdeplot(passing_23, shade=True, label='Passing (2023)', color='brown', bw_adjust=1.2)

# Titles and labels
plt.xlabel('Percentage', fontsize = 15, labelpad=15, fontweight = 'bold')
plt.ylabel('Density', fontsize = 15, labelpad=15, fontweight = 'bold')
plt.title('Rushing and Passing Play % (2003, 2013, 2023)')

# Add legend
plt.legend(loc='upper right')


## Save the figure as an image file (e.g., PNG or JPG)
plt.savefig('Figures/play_percentage_distribution_20years.png', format='png', dpi=300)





## Figure 2 'Win Percentage vs Rushing Play Percentage in 2023' ##

# Create a dictionary of NFL teams and their primary colors in hex
team_colors = {
    'Baltimore': '#241773', 'Kansas City': '#E31837', 'San Francisco': '#AA0000', 'Detroit': '#0076B6',
    'Dallas': '#003594', 'Buffalo': '#00338D', 'Cleveland': '#311D00', 'Miami': '#008E97',
    'Philadelphia': '#004C54', 'Houston': '#03202F', 'LA Rams': '#003594', 'Pittsburgh': '#FFB612',
    'Cincinnati': '#FB4F14', 'New Orleans': '#D3BC8D', 'Seattle': '#002244', 'Jacksonville': '#006778',
    'Indianapolis': '#002C5F', 'Tampa Bay': '#D50A0A', 'Green Bay': '#203731', 'Denver': '#FB4F14',
    'Las Vegas': '#000000', 'Minnesota': '#4F2683', 'Atlanta': '#A71930', 'Chicago': '#0B162A',
    'NY Jets': '#125740', 'NY Giants': '#0B2265', 'Tennessee': '#4B92DB', 'LA Chargers': '#0073CF',
    'Washington': '#5A1414', 'New England': '#002244', 'Arizona': '#97233F', 'Carolina': '#0085CA'
}

# Calculate correlation coefficients for each year
correlation_2003 = df['Rushing Play % (2003)'].corr(df['Win % (2003)'])
correlation_2013 = df['Rushing Play % (2013)'].corr(df['Win % (2013)'])
correlation_2023 = df['Rushing Play % (2023)'].corr(df['Win % (2023)'])

print(f'Correlation (2003): {correlation_2003:.2f}')
print(f'Correlation (2013): {correlation_2013:.2f}')
print(f'Correlation (2023): {correlation_2023:.2f}')

# Function to create scatter plot for a given year
def plot_scatter(year):
    plt.figure(figsize=(12, 6))
    
    # Select data for the year
    x_data = df[f'Rushing Play % ({year})']
    y_data = df[f'Win % ({year})']
    teams = df['Team']
    
    # Map each team to its color
    colors = [team_colors[team] for team in teams]

    # Create the scatter plot
    scatter = plt.scatter(x_data, y_data, color=colors, s=100)
    
    
    # Add titles and labels
    plt.title(f'Win Percentage vs Rushing Play Percentage in {year}', fontsize=14)
    plt.xlabel('Rushing Play Percentage (%)', fontsize=12, fontweight = 'bold')
    plt.ylabel('Win Percentage (%)', fontsize=12, fontweight = 'bold')


    # Create custom legend with team names and their colors
    handles = [plt.Line2D([0], [0], marker='o', color=color, markersize=10, linestyle='', label=team)
               for team, color in team_colors.items()]

    # Position the legend at the bottom with multiple columns
    plt.legend(handles=handles, bbox_to_anchor=(0.5, -0.2), loc='upper center', ncol=6, title='NFL Teams')

    # Calculate and display the correlation coefficient
    correlation = x_data.corr(y_data)

    # Adjust layout
    plt.tight_layout()

    # Display the correlation coefficient on the plot
    plt.text(0.05, 0.95, f'Correlation: {correlation:.2f}', fontsize=12, color='black',
             transform=plt.gca().transAxes, verticalalignment='top', horizontalalignment='left')


    # Save the plot
    plt.savefig(f'Figures/win_vs_rush ({year}).png', format='png', dpi=300)


# Generate scatter plots for each year
for year in [2003, 2013, 2023]:
    plot_scatter(year)





## Figure 3 'Rushing vs Passing Tocuhdown % (2023) ##

# Select only the relevant columns
td_data = df[['Rushing Touchdowns %', 'Passing Touchdown %']]

# Convert to long format for easier plotting
td_long = td_data.melt(var_name='TD Type', value_name='Touchdown %')

# Rename the x-axis values
td_long['TD Type'] = td_long['TD Type'].replace({
    'Rushing Touchdowns %': 'Rushing',
    'Passing Touchdown %': 'Passing'
})

# Plot the box plot
plt.figure(figsize=(8, 6))
sns.boxplot(x='TD Type', y='Touchdown %', data=td_long, palette={'Rushing': 'lightgreen', 'Passing': 'skyblue'})

# Add titles and labels
plt.title('Rushing vs Passing Touchdown % (2023)')
plt.xlabel('Touchdown Type', labelpad=15, fontweight = 'bold')
plt.ylabel('Percentage', labelpad=15, fontweight = 'bold')

plt.savefig('Figures/touchdown_type_percentages.png', format='png', dpi=300)





## Figure 4 'Pass Completion %'s' ##

# Select only the columns with completion percentages
completion_columns = ['Completion % (2003)', 'Completion % (2013)', 'Completion % (2023)']

# Create a DataFrame with only these columns
completion_data = df[completion_columns]

# Convert the data to long format for easier plotting
completion_long = completion_data.melt(var_name='Year', value_name='Completion %')

# Rename the x-axis values
completion_long['Year'] = completion_long['Year'].replace({
    'Completion % (2003)': '2003',
    'Completion % (2013)': '2013',
    'Completion % (2023)': '2023'
})

# Define a custom color palette for each year
custom_palette = {"2003": "skyblue", "2013": "lightgreen", "2023": "#FF9999"}

# Plot box plots for each year's completion percentage on the same axis
plt.figure(figsize=(10, 6))
sns.boxplot(x='Year', y='Completion %', data=completion_long, palette = custom_palette)
plt.title('Pass Completion %`s')
plt.xlabel('Year', fontsize = 10, labelpad=15, fontweight = 'bold')
plt.ylabel('Completion Percentage', fontsize = 10, fontweight = 'bold')

plt.savefig('Figures/pass_completion_percentages.png', format='png', dpi=300)
